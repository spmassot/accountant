import pandas as pd
from xlsxmetadata.metadata import get_sheet_names
from io import BytesIO
from openpyxl import load_workbook
import re


def prep_file(ifile, file_type):
    return {
        'ups': prep_ups_file,
        'givens': prep_givens_file,
        'freight': prep_freight_report
    }.get(file_type)(ifile)


def prep_ups_file(ups_file):
    wb = load_workbook(
        filename=BytesIO(ups_file.read()),
        read_only=True,
        data_only=True
    )
    sought = re.compile(r'^\s*accruals*\s*$', re.I)
    matched = [x for x in wb.sheetnames if sought.match(x)]
    if not matched:
        raise ValueError(f'There is no Accruals sheet; sheetnames: {wb.sheetnames}')
    rows = wb[matched[0]].iter_rows(values_only=True)

    keep_columns = ups_columns()
    header = next(rows)
    return [
        {
            keep_columns[header[i]]: itm for i, itm in enumerate(row)
            if header[i] in keep_columns
        }
        for row in rows
    ][:-1]


def ups_columns():
    return {
        'Recipient Number': 'recipient_number',
        'Account Number': 'account_number',
        'Invoice Date': 'invoice_date',
        'Invoice Number': 'invoice_number',
        'Invoice Amount': 'invoice_amount',
        'Transaction Date': 'transaction_date',
        'Lead Shipment Number': 'lead_shipment_number',
        'Shipment Reference Number 1': 'shipment_reference_number_1',
        'Shipment Reference Number 2': 'shipment_reference_number_2',
        'Tracking Number': 'tracking_number',
        'Package Reference Number 1': 'package_reference_number_1',
        'Package Reference Number 2': 'package_reference_number_2',
        'Package Reference Number 3': 'package_reference_number_3',
        'Package Reference Number 4': 'package_reference_number_4',
        'Package Reference Number 5': 'package_reference_number_5',
        'Net Amount': 'net_amount',
    }


def prep_givens_file(givens_file):
    sought = ['Brampton', 'Chesapeake','Plainfield','Reno','Hope']
    adj_sought = [x + ' Adjustments' for x in sought]
    keep_columns = givens_columns()

    frames = pd.read_excel(givens_file, sheet_name=None, converters={"ORDER #'s":str})

    flat_names = [x for x in frames if x in sought]
    adj_names = [x for x in frames if x in adj_sought]
    all_names = flat_names + adj_names

    frames = {k: v for k, v in frames.items() if k in all_names}

    flat_frames = []
    adj_frames = []

    for name, df in frames.items():
        df.columns = df.iloc[5]
        df = df.drop(labels=range(6), axis=0)

        if name in flat_names:
            df = df.dropna(thresh=12)
            df['location'] = name
            df = df[[*list(keep_columns.keys())]]
            df = df.rename(keep_columns, axis='columns')
            flat_frames.extend(df.to_dict(orient='records'))

        elif name in adj_names:
            df = df.dropna(thresh=4)
            df['location'] = name.replace('Adjustments','').strip()
            df = df[[*list(keep_columns.keys())]]
            df = df.rename(keep_columns, axis='columns')
            adj_frames.extend(df.to_dict(orient='records'))

    return (flat_frames, adj_frames)


def givens_columns():
    return {
        'LOAD NUMBER': 'load_number',
        "ORDER #'s": 'order_numbers',
        'SHIP DATE': 'ship_date',
        'CARRIER': 'carrier',
        'TOTAL': 'total'
    }


def prep_freight_report(freight_report):

    keep_columns = freight_report_columns()
    frames = pd.read_excel(freight_report, sheet_name=None)

    records = []
    for k, df in frames.items():
        df = df[[*list(keep_columns.keys())]]
        df.rename(keep_columns, axis='columns')
        records.append(df)

    return pd.concat(records).to_dict(orient='records')


def freight_report_columns():
    return {
        'MM': 'month',
        'YY': 'year',
        'Pick Ticket': 'pick_ticket',
        'Order': 'order_number',
        'Doc#': 'document_number',
        'Sales': 'sales',
        'Cost': 'cost',
        'Freight Sales': 'freight_sales',
        'Freight Cost': 'freight_cost',
        'Carrier#': 'carrier_number',
        'Name': 'carrier',
        'Company': 'company',
        'Tracking Number': 'tracking_number',
        'Original Pick Ticket 1': 'original_pick_ticket'
    }
