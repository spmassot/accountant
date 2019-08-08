from openpyxl import load_workbook
from io import BytesIO
from src.preppers import Prepper
import re


class UpsPrepper(Prepper):
    column_mapping = {
        'Recipient Number': 'recipient_number',
        'Account Number': 'account_number',
        'Invoice Date': 'invoice_date',
        'Invoice Number': 'invoice_number',
        'Invoice Amount': 'invoice_amount',
        'Transaction Date': 'transaction_date',
        'Lead Shipment Number': 'lead_shipment_number',
        'Shipment Reference Number 1': 'shipment_reference_number_1',
        'Shipment Reference Number 2': 'shipment_reference_number_2',
        'Tracking Number': 'tracking_number',
        'Package Reference Number 1': 'package_reference_number_1',
        'Package Reference Number 2': 'package_reference_number_2',
        'Package Reference Number 3': 'package_reference_number_3',
        'Package Reference Number 4': 'package_reference_number_4',
        'Package Reference Number 5': 'package_reference_number_5',
        'Net Amount': 'net_amount',
    }

    def prep_file(self):
        wb = load_workbook(
            filename=BytesIO(self.ifile.read()),
            read_only=True,
            data_only=True
        )
        sought = re.compile(r'^\s*accruals*\s*$', re.I)
        matched = [x for x in wb.sheetnames if sought.match(x)]
        if not matched:
            raise ValueError(f'There is no Accruals sheet; sheetnames: {wb.sheetnames}')
        rows = wb[matched[0]].iter_rows(values_only=True)

        header = next(rows)
        return [{
            self.column_mapping[header[i]]: itm for i, itm in enumerate(row)
            if header[i] in self.column_mapping
        } for row in rows][:-1]
